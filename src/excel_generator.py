"""Генератор XLSX файлов для заказов карточек MTG."""

import logging
from decimal import Decimal
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models import Card

# Настройка логирования
logger = logging.getLogger(__name__)


def generate_excel(cards: List[Card], output_path: str, use_formulas: bool = True) -> None:
    """
    Генерирует Excel файл из списка карточек MTG.
    
    Создает новый Excel файл (Workbook) с одним листом "Лист1", содержащий:
    - Заголовки в первой строке
    - Данные карт с форматированием
    - Гиперссылки в колонке "Ссылка"
    - Формулы или вычисленные значения в колонке "Итого"
    - Денежный формат для цен
    
    Args:
        cards (List[Card]): Список объектов Card для экспорта.
        output_path (str): Путь к файлу для сохранения (например, "order.xlsx").
        use_formulas (bool): Если True, использует формулы (=A*G) для итоговых сумм.
                            Если False, записывает вычисленные значения. По умолчанию True.
    
    Raises:
        ValueError: Если список карт пустой.
        OSError: Если возникла ошибка при записи файла.
    
    Example:
        >>> cards = [Card(...), Card(...)]
        >>> generate_excel(cards, "my_order.xlsx")
    """
    # Валидация входных данных
    if not cards:
        logger.error("Попытка создать Excel файл с пустым списком карт")
        raise ValueError("Список карт не может быть пустым")
    
    logger.info(f"Начинается генерация Excel файла: {output_path}")
    logger.info(f"Количество карт: {len(cards)}, использование формул: {use_formulas}")
    
    try:
        # Создание нового Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"
        
        # Заголовки таблицы
        headers = [
            "Количество",
            "Название карты",
            "Ссылка",
            "Фойл",
            "Состояние",
            "Издание",
            "Цена за штуку (USD)",
            "Итого:"
        ]
        ws.append(headers)
        logger.debug(f"Заголовки добавлены: {headers}")
        
        # Добавление данных карт
        for idx, card in enumerate(cards, start=2):
            # Получаем данные карты через метод to_excel_row()
            row_data = card.to_excel_row()
            
            # Подготовка данных для записи
            # row_data содержит: [quantity, name, url, foil_text, condition, edition, price_per_unit, total_price]
            # Нам нужно заменить последнее значение (total_price) на формулу, если use_formulas=True
            if use_formulas:
                # Заменяем итоговую сумму на формулу
                row_to_write = row_data[:7] + [f"=A{idx}*G{idx}"]
            else:
                # Используем вычисленное значение, конвертируем Decimal в float
                row_to_write = row_data[:7] + [float(row_data[7])]
            
            # Конвертируем Decimal в float для цены за штуку (индекс 6)
            row_to_write[6] = float(row_to_write[6])
            
            ws.append(row_to_write)
            
            # Создание гиперссылки в колонке C (Ссылка)
            cell_ref = f'C{idx}'
            ws[cell_ref].hyperlink = card.url
            ws[cell_ref].style = 'Hyperlink'
            
            logger.debug(f"Строка {idx}: добавлена карта '{card.name}'")
        
        # Форматирование worksheet
        _format_worksheet(ws)
        
        # Сохранение файла
        wb.save(output_path)
        logger.info(f"Excel файл успешно сохранен: {output_path}")
        
    except OSError as e:
        logger.error(f"Ошибка записи файла {output_path}: {e}")
        raise OSError(f"Не удалось записать файл {output_path}: {e}") from e
    except Exception as e:
        logger.error(f"Неожиданная ошибка при генерации Excel: {e}")
        raise


def _format_worksheet(ws: Worksheet) -> None:
    """
    Применяет форматирование к worksheet.
    
    Выполняет следующие операции форматирования:
    - Делает заголовок (первая строка) жирным шрифтом
    - Устанавливает автоширину для всех колонок
    - Выравнивает числовые колонки (A, G, H) по правому краю
    - Применяет денежный формат ($#,##0.00) к ценам (колонки G и H)
    - Центрирует колонку "Фойл" (D)
    
    Args:
        ws (Worksheet): Worksheet для форматирования.
    
    Example:
        >>> _format_worksheet(worksheet)
    """
    logger.debug("Начинается форматирование worksheet")
    
    # 1. Жирный шрифт для заголовков (первая строка)
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    logger.debug("Заголовки отформатированы жирным шрифтом")
    
    # 2. Форматирование колонок
    # A - Количество (число, выравнивание по правому краю)
    # B - Название карты (текст, по левому краю - по умолчанию)
    # C - Ссылка (гиперссылка, по левому краю)
    # D - Фойл (текст, по центру)
    # E - Состояние (текст, по левому краю)
    # F - Издание (текст, по левому краю)
    # G - Цена за штуку (денежный формат, по правому краю)
    # H - Итого (денежный формат, по правому краю)
    
    right_align = Alignment(horizontal='right')
    center_align = Alignment(horizontal='center')
    money_format = '$#,##0.00'
    
    # Колонка A: Количество - выравнивание по правому краю
    for cell in ws['A'][1:]:  # Пропускаем заголовок
        cell.alignment = right_align
    
    # Колонка D: Фойл - выравнивание по центру
    for cell in ws['D'][1:]:  # Пропускаем заголовок
        cell.alignment = center_align
    
    # Колонка G: Цена за штуку - денежный формат и выравнивание по правому краю
    for cell in ws['G'][1:]:  # Пропускаем заголовок
        cell.number_format = money_format
        cell.alignment = right_align
    
    # Колонка H: Итого - денежный формат и выравнивание по правому краю
    for cell in ws['H'][1:]:  # Пропускаем заголовок
        cell.number_format = money_format
        cell.alignment = right_align
    
    logger.debug("Форматирование числовых колонок и выравнивание применено")
    
    # 3. Автоширина колонок
    # Вычисляем оптимальную ширину для каждой колонки на основе содержимого
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        
        # Находим максимальную длину контента в колонке
        max_length = 0
        for cell in column_cells:
            try:
                # Учитываем длину значения ячейки
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Устанавливаем ширину колонки с небольшим запасом
        # Минимальная ширина: 10, максимальная: 50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    logger.debug("Автоширина колонок установлена")
    logger.info("Форматирование worksheet завершено")