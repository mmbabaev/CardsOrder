"""Unit tests для генератора Excel файлов."""

import os
import tempfile
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Font

from src.excel_generator import generate_excel, _format_worksheet
from src.models import Card


# Фикстуры для тестовых данных
@pytest.fixture
def sample_cards():
    """Создает список тестовых карт для использования в тестах."""
    return [
        Card(
            quantity=2,
            name="Lightning Bolt",
            url="https://www.cardkingdom.com/mtg/alpha/lightning-bolt",
            is_foil=False,
            condition="NM",
            edition="Alpha",
            edition_code="LEA",
            price_per_unit=Decimal("99.99"),
            total_price=Decimal("199.98"),
            rarity="C",
            variation=None
        ),
        Card(
            quantity=1,
            name="Black Lotus",
            url="https://www.cardkingdom.com/mtg/alpha/black-lotus",
            is_foil=False,
            condition="EX",
            edition="Alpha",
            edition_code="LEA",
            price_per_unit=Decimal("25000.00"),
            total_price=Decimal("25000.00"),
            rarity="R",
            variation=None
        ),
        Card(
            quantity=4,
            name="Flare of Denial",
            url="https://www.cardkingdom.com/mtg/modern-horizons-3/flare-of-denial",
            is_foil=True,
            condition="NM",
            edition="Modern Horizons 3",
            edition_code="MH3",
            price_per_unit=Decimal("59.99"),
            total_price=Decimal("239.96"),
            rarity="R",
            variation="0400 - Retro Frame"
        )
    ]


@pytest.fixture
def temp_excel_file():
    """Создает временный файл Excel для тестов."""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False) as f:
        temp_path = f.name
    yield temp_path
    # Cleanup
    if os.path.exists(temp_path):
        os.unlink(temp_path)


class TestGenerateExcel:
    """Тесты для функции generate_excel."""
    
    def test_creates_file_successfully(self, sample_cards, temp_excel_file):
        """Тест: файл Excel создается успешно."""
        generate_excel(sample_cards, temp_excel_file)
        
        assert os.path.exists(temp_excel_file), "Excel файл должен быть создан"
        assert os.path.getsize(temp_excel_file) > 0, "Файл не должен быть пустым"
    
    def test_creates_correct_worksheet_name(self, sample_cards, temp_excel_file):
        """Тест: создается worksheet с правильным именем 'Лист1'."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        assert wb.active.title == "Лист1", "Название листа должно быть 'Лист1'"
        wb.close()
    
    def test_headers_are_correct(self, sample_cards, temp_excel_file):
        """Тест: заголовки таблицы корректны."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        expected_headers = [
            "Количество",
            "Название карты",
            "Ссылка",
            "Фойл",
            "Состояние",
            "Издание",
            "Цена за штуку (USD)",
            "Итого:"
        ]
        
        actual_headers = [cell.value for cell in ws[1]]
        assert actual_headers == expected_headers, "Заголовки должны совпадать с ожидаемыми"
        wb.close()
    
    def test_headers_are_bold(self, sample_cards, temp_excel_file):
        """Тест: заголовки имеют жирный шрифт."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        for cell in ws[1]:
            assert cell.font.bold is True, f"Заголовок '{cell.value}' должен быть жирным"
        wb.close()
    
    def test_correct_number_of_rows(self, sample_cards, temp_excel_file):
        """Тест: количество строк соответствует количеству карт + заголовок."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # +1 для заголовка
        expected_rows = len(sample_cards) + 1
        assert ws.max_row == expected_rows, f"Должно быть {expected_rows} строк (включая заголовок)"
        wb.close()
    
    def test_data_correctness(self, sample_cards, temp_excel_file):
        """Тест: данные в ячейках корректны."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Проверяем первую карту (Lightning Bolt)
        assert ws['A2'].value == 2, "Количество должно быть 2"
        assert ws['B2'].value == "Lightning Bolt", "Название должно совпадать"
        assert ws['C2'].value == "https://www.cardkingdom.com/mtg/alpha/lightning-bolt", "URL должен совпадать"
        assert ws['D2'].value == "НЕТ", "Не фойловая карта должна иметь значение 'НЕТ'"
        assert ws['E2'].value == "NM", "Состояние должно быть NM"
        assert ws['F2'].value == "LEA", "Должен использоваться код издания LEA"
        assert ws['G2'].value == 99.99, "Цена за штуку должна быть 99.99"
        
        # Проверяем вторую карту (Black Lotus)
        assert ws['A3'].value == 1, "Количество должно быть 1"
        assert ws['B3'].value == "Black Lotus", "Название должно совпадать"
        
        # Проверяем третью карту (Flare of Denial - foil)
        assert ws['D4'].value == "ДА", "Фойловая карта должна иметь значение 'ДА'"
        assert ws['F4'].value == "MH3", "Должен использоваться код издания MH3"
        
        wb.close()
    
    def test_hyperlinks_in_link_column(self, sample_cards, temp_excel_file):
        """Тест: колонка 'Ссылка' содержит гиперссылки."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Проверяем гиперссылки для каждой карты
        for idx, card in enumerate(sample_cards, start=2):
            cell = ws[f'C{idx}']
            assert cell.hyperlink is not None, f"Ячейка C{idx} должна содержать гиперссылку"
            assert cell.hyperlink.target == card.url, f"URL гиперссылки должен совпадать с URL карты"
            assert cell.style == 'Hyperlink', f"Ячейка C{idx} должна иметь стиль Hyperlink"
        
        wb.close()
    
    def test_formulas_in_total_column(self, sample_cards, temp_excel_file):
        """Тест: колонка 'Итого' содержит формулы при use_formulas=True."""
        generate_excel(sample_cards, temp_excel_file, use_formulas=True)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Проверяем формулы для каждой карты
        for idx in range(2, len(sample_cards) + 2):
            cell = ws[f'H{idx}']
            expected_formula = f"=A{idx}*G{idx}"
            # openpyxl сохраняет формулу с знаком "="
            assert cell.value == expected_formula, f"Ячейка H{idx} должна содержать формулу {expected_formula}"
        
        wb.close()
    
    def test_calculated_values_without_formulas(self, sample_cards, temp_excel_file):
        """Тест: колонка 'Итого' содержит вычисленные значения при use_formulas=False."""
        generate_excel(sample_cards, temp_excel_file, use_formulas=False)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Проверяем вычисленные значения
        assert ws['H2'].value == 199.98, "Итого для Lightning Bolt должно быть 199.98"
        assert ws['H3'].value == 25000.00, "Итого для Black Lotus должно быть 25000.00"
        assert ws['H4'].value == 239.96, "Итого для Flare of Denial должно быть 239.96"
        
        wb.close()
    
    def test_money_format_for_prices(self, sample_cards, temp_excel_file):
        """Тест: денежный формат применен к ценам в колонках G и H."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        money_format = '$#,##0.00'
        
        # Проверяем формат для колонки G (Цена за штуку)
        for idx in range(2, len(sample_cards) + 2):
            cell = ws[f'G{idx}']
            assert cell.number_format == money_format, f"Ячейка G{idx} должна иметь денежный формат"
        
        # Проверяем формат для колонки H (Итого)
        for idx in range(2, len(sample_cards) + 2):
            cell = ws[f'H{idx}']
            assert cell.number_format == money_format, f"Ячейка H{idx} должна иметь денежный формат"
        
        wb.close()
    
    def test_alignment_for_numeric_columns(self, sample_cards, temp_excel_file):
        """Тест: выравнивание числовых колонок по правому краю."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Колонка A (Количество) - по правому краю
        for idx in range(2, len(sample_cards) + 2):
            assert ws[f'A{idx}'].alignment.horizontal == 'right', f"Ячейка A{idx} должна быть выровнена по правому краю"
        
        # Колонка G (Цена) - по правому краю
        for idx in range(2, len(sample_cards) + 2):
            assert ws[f'G{idx}'].alignment.horizontal == 'right', f"Ячейка G{idx} должна быть выровнена по правому краю"
        
        # Колонка H (Итого) - по правому краю
        for idx in range(2, len(sample_cards) + 2):
            assert ws[f'H{idx}'].alignment.horizontal == 'right', f"Ячейка H{idx} должна быть выровнена по правому краю"
        
        wb.close()
    
    def test_foil_column_centered(self, sample_cards, temp_excel_file):
        """Тест: колонка 'Фойл' выровнена по центру."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Колонка D (Фойл) - по центру
        for idx in range(2, len(sample_cards) + 2):
            assert ws[f'D{idx}'].alignment.horizontal == 'center', f"Ячейка D{idx} должна быть выровнена по центру"
        
        wb.close()
    
    def test_raises_error_on_empty_card_list(self, temp_excel_file):
        """Тест: выбрасывается ValueError при пустом списке карт."""
        with pytest.raises(ValueError, match="Список карт не может быть пустым"):
            generate_excel([], temp_excel_file)
    
    def test_raises_error_on_invalid_path(self, sample_cards):
        """Тест: выбрасывается OSError при невалидном пути к файлу."""
        invalid_path = "/nonexistent/directory/file.xlsx"
        
        with pytest.raises(OSError):
            generate_excel(sample_cards, invalid_path)
    
    def test_column_widths_are_set(self, sample_cards, temp_excel_file):
        """Тест: ширина колонок установлена (автоширина)."""
        generate_excel(sample_cards, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Проверяем, что ширина колонок установлена и находится в разумных пределах
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            width = ws.column_dimensions[col_letter].width
            assert width >= 10, f"Ширина колонки {col_letter} должна быть не меньше 10"
            assert width <= 50, f"Ширина колонки {col_letter} должна быть не больше 50"
        
        wb.close()
    
    def test_handles_cards_without_edition_code(self, temp_excel_file):
        """Тест: корректно обрабатывает карты без кода издания."""
        cards_without_code = [
            Card(
                quantity=1,
                name="Test Card",
                url="https://www.cardkingdom.com/mtg/test/test-card",
                is_foil=False,
                condition="NM",
                edition="Test Edition Name",
                edition_code=None,  # Нет кода издания
                price_per_unit=Decimal("1.00"),
                total_price=Decimal("1.00"),
                rarity="C",
                variation=None
            )
        ]
        
        generate_excel(cards_without_code, temp_excel_file)
        
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Должно использоваться полное название издания
        assert ws['F2'].value == "Test Edition Name", "Должно использоваться полное название издания"
        wb.close()


class TestFormatWorksheet:
    """Тесты для функции _format_worksheet."""
    
    def test_format_worksheet_applies_formatting(self, sample_cards, temp_excel_file):
        """Тест: _format_worksheet применяет все необходимое форматирование."""
        # Создаем простой Excel файл без форматирования
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"
        
        headers = ["Количество", "Название карты", "Ссылка", "Фойл", 
                   "Состояние", "Издание", "Цена за штуку (USD)", "Итого:"]
        ws.append(headers)
        ws.append([1, "Test", "http://test.com", "НЕТ", "NM", "TEST", 1.00, 1.00])
        
        # Применяем форматирование
        _format_worksheet(ws)
        
        # Проверяем жирный шрифт заголовков
        for cell in ws[1]:
            assert cell.font.bold is True, "Заголовки должны быть жирными"
        
        # Проверяем выравнивание
        assert ws['A2'].alignment.horizontal == 'right', "Количество должно быть по правому краю"
        assert ws['D2'].alignment.horizontal == 'center', "Фойл должен быть по центру"
        assert ws['G2'].alignment.horizontal == 'right', "Цена должна быть по правому краю"
        assert ws['H2'].alignment.horizontal == 'right', "Итого должно быть по правому краю"
        
        # Проверяем денежный формат
        money_format = '$#,##0.00'
        assert ws['G2'].number_format == money_format, "Цена должна иметь денежный формат"
        assert ws['H2'].number_format == money_format, "Итого должно иметь денежный формат"
        
        wb.close()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])