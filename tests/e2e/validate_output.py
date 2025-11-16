#!/usr/bin/env python3
"""
Script to validate the test_output.xlsx file
"""
import openpyxl
from openpyxl.styles import Font
import json

def validate_excel(filename):
    """Validate the Excel file and return a report"""
    print(f"Opening file: {filename}")
    # Load workbook twice: once for formulas, once for data
    wb = openpyxl.load_workbook(filename)
    wb_data = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    ws_data = wb_data.active
    
    report = {
        "filename": filename,
        "tests": {},
        "issues": [],
        "statistics": {}
    }
    
    # Test 1: Check row count
    print("\n=== Test 1: Row Count ===")
    max_row = ws.max_row
    expected_rows = 22  # 1 header + 21 cards
    test_passed = max_row == expected_rows
    report["tests"]["row_count"] = {
        "expected": expected_rows,
        "actual": max_row,
        "passed": test_passed
    }
    print(f"Expected: {expected_rows}, Actual: {max_row}, Passed: {test_passed}")
    if not test_passed:
        report["issues"].append(f"Row count mismatch: expected {expected_rows}, got {max_row}")
    
    # Test 2: Check column headers
    print("\n=== Test 2: Column Headers ===")
    expected_headers = [
        'Количество', 'Название карты', 'Ссылка', 'Фойл', 
        'Состояние', 'Издание', 'Цена за штуку (USD)', 'Итого:'
    ]
    actual_headers = [cell.value for cell in ws[1]]
    headers_match = actual_headers == expected_headers
    report["tests"]["headers"] = {
        "expected": expected_headers,
        "actual": actual_headers,
        "passed": headers_match
    }
    print(f"Expected headers: {expected_headers}")
    print(f"Actual headers: {actual_headers}")
    print(f"Passed: {headers_match}")
    if not headers_match:
        report["issues"].append(f"Headers mismatch")
    
    # Test 3: Check header formatting (bold)
    print("\n=== Test 3: Header Formatting ===")
    headers_bold = all(cell.font.bold for cell in ws[1])
    report["tests"]["header_formatting"] = {
        "bold": headers_bold,
        "passed": headers_bold
    }
    print(f"Headers are bold: {headers_bold}")
    if not headers_bold:
        report["issues"].append("Headers are not bold")
    
    # Test 4: Check formulas in 'Итого:' column (column H)
    print("\n=== Test 4: Formulas in 'Итого:' Column ===")
    formulas_found = 0
    formula_issues = []
    for row_idx in range(2, max_row + 1):
        cell = ws[f'H{row_idx}']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formulas_found += 1
            # Check if it's the correct formula pattern
            expected_pattern = f'=A{row_idx}*G{row_idx}'
            if cell.value != expected_pattern:
                formula_issues.append(f"Row {row_idx}: expected '{expected_pattern}', got '{cell.value}'")
    
    formulas_correct = formulas_found == (max_row - 1) and len(formula_issues) == 0
    report["tests"]["formulas"] = {
        "expected_count": max_row - 1,
        "actual_count": formulas_found,
        "issues": formula_issues,
        "passed": formulas_correct
    }
    print(f"Formulas found: {formulas_found}/{max_row - 1}")
    if formula_issues:
        print(f"Formula issues: {formula_issues}")
    print(f"Passed: {formulas_correct}")
    if not formulas_correct:
        report["issues"].append(f"Formula issues found: {len(formula_issues)}")
    
    # Test 5: Check hyperlinks in 'Ссылка' column (column C)
    print("\n=== Test 5: Hyperlinks ===")
    hyperlinks_found = 0
    for row_idx in range(2, max_row + 1):
        cell = ws[f'C{row_idx}']
        if cell.hyperlink:
            hyperlinks_found += 1
    
    hyperlinks_correct = hyperlinks_found == (max_row - 1)
    report["tests"]["hyperlinks"] = {
        "expected_count": max_row - 1,
        "actual_count": hyperlinks_found,
        "passed": hyperlinks_correct
    }
    print(f"Hyperlinks found: {hyperlinks_found}/{max_row - 1}")
    print(f"Passed: {hyperlinks_correct}")
    if not hyperlinks_correct:
        report["issues"].append(f"Hyperlinks count mismatch: expected {max_row - 1}, got {hyperlinks_found}")
    
    # Test 6: Check data types
    print("\n=== Test 6: Data Types ===")
    quantity_issues = []
    price_issues = []
    
    for row_idx in range(2, max_row + 1):
        # Check quantity (column A) is numeric
        quantity_cell = ws[f'A{row_idx}']
        if not isinstance(quantity_cell.value, (int, float)):
            quantity_issues.append(f"Row {row_idx}: {quantity_cell.value} is not numeric")
        
        # Check price (column G) is numeric
        price_cell = ws[f'G{row_idx}']
        if not isinstance(price_cell.value, (int, float)):
            price_issues.append(f"Row {row_idx}: {price_cell.value} is not numeric")
    
    data_types_correct = len(quantity_issues) == 0 and len(price_issues) == 0
    report["tests"]["data_types"] = {
        "quantity_issues": quantity_issues,
        "price_issues": price_issues,
        "passed": data_types_correct
    }
    print(f"Quantity issues: {len(quantity_issues)}")
    print(f"Price issues: {len(price_issues)}")
    print(f"Passed: {data_types_correct}")
    if not data_types_correct:
        report["issues"].extend(quantity_issues + price_issues)
    
    # Statistics
    print("\n=== Statistics ===")
    total_quantity = sum(ws[f'A{row_idx}'].value for row_idx in range(2, max_row + 1))
    # Manually calculate total sum since formulas aren't evaluated by openpyxl
    total_sum = sum(ws[f'A{row_idx}'].value * ws[f'G{row_idx}'].value for row_idx in range(2, max_row + 1))
    foil_count = sum(1 for row_idx in range(2, max_row + 1) if ws[f'D{row_idx}'].value == 'Да')
    
    # Also count НЕТ foils for reporting
    non_foil_count = sum(1 for row_idx in range(2, max_row + 1) if ws[f'D{row_idx}'].value == 'НЕТ')
    
    report["statistics"] = {
        "total_cards": max_row - 1,
        "total_quantity": total_quantity,
        "total_sum": round(total_sum, 2),
        "foil_count": foil_count
    }
    
    print(f"Total cards: {max_row - 1}")
    print(f"Total quantity: {total_quantity}")
    print(f"Total sum: ${total_sum:.2f}")
    print(f"Foil cards: {foil_count}")
    
    # Overall result
    all_tests_passed = all(test["passed"] for test in report["tests"].values())
    report["overall_passed"] = all_tests_passed
    
    print("\n" + "=" * 50)
    print(f"Overall Test Result: {'PASSED' if all_tests_passed else 'FAILED'}")
    print(f"Issues found: {len(report['issues'])}")
    print("=" * 50)
    
    return report

if __name__ == "__main__":
    report = validate_excel("test_output.xlsx")
    
    # Save report as JSON
    with open("validation_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    
    print("\nValidation report saved to: validation_report.json")